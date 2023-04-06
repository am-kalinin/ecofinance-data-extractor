#!/usr/bin/python3

PDL_Zero = []
PDL_Repeated = []
MINI = []
MINI56 = []

import openpyxl

from openpyxl import load_workbook

def readcsv(product):
    is_work=0
    result = []
    filename='/tmp/recovery_'+product+'.csv'
    f=open(filename)
    for row in f:
        l = row.split(',')
        if (is_work==0 and l[0]!='FROM TOTAL OPEN'):
            continue
        if (is_work==0 and l[0]=='FROM TOTAL OPEN'):
            is_work = 1
        if (is_work == 1 and l[0]!= 'J. 121-150 dpd'):
            if (l[0] == 'A. 1-5 dpd' or l[0] == 'B. 6-10 dpd' or l[0] == 'C. 11-15 dpd' or l[0] == 'D. 16-30 dpd' or l[0] == 'E. 31-45 dpd' or l[0] == 'F. 45-60 dpd' or l[0] == 'H. 61-90 dpd' or l[0] == 'H. 61-90 dpd' or l[0] == 'I. 91-120 dpd'):
                result.append(l[5])
            elif (l[0] == 'J. 121-150 dpd'):
                is_work = 2
        if (is_work == 2):
            pass
    return (result)

PDL_Zero = readcsv('pdl_zero')
PDL_Repeated = readcsv('pdl_repeated')
MINI = readcsv('mini')
MINI56 = readcsv('mini56')

print ('<table>') #<tr><td>PDLZ</td><td>PDLR</td><td>MINI</td><td>MINI56</td></tr>')
for a in range(0,8):
    print (f'<tr><td>{float(PDL_Zero[a])*100}%</td><td>{float(PDL_Repeated[a])*100}%</td><td>{float(MINI[a])*100}%</td><td>{float(MINI56[a])*100}%</td></tr>')
print ('</table>')

wb=load_workbook(filename='targers.xlsx')
ws=wb['Sheet1']

def get_target (col):
    targ = []
    for a in range (2,4):
        place = col+str(a)
        targ.append(float(ws[place].value.split('%')[0])/100)
    return (targ)

target_mini = get_target('B')
target_pdl_zero = get_target('C')
target_pdl_repeated = get_target('D')
target_mini56 = get_target('E')

config_of_mess=load_workbook(filename='config.xlsx')
ws=config_of_mess['data']

def real_recovery (rate, target):
    result = []
    for a in range(2):
        recovery = float(rate[a])/float(target[a])*100
        result.append(recovery)
    return (result)

def write_recovery_to_file (result, col):
    ws[col+str(2)].value = str(result[0])+'%'
    ws[col+str(3)].value = str(result[1])+'%'

write_recovery_to_file(real_recovery(PDL_Zero, target_pdl_zero), 'D')
write_recovery_to_file(real_recovery(PDL_Repeated, target_pdl_repeated), 'E')
write_recovery_to_file(real_recovery(MINI, target_mini), 'F')
write_recovery_to_file(real_recovery(MINI56, target_mini56), 'G')

config_of_mess.save('config.xlsx')
