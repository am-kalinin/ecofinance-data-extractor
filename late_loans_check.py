#!/usr/bin/python3

import sys, csv, openpyxl

from openpyxl import load_workbook

class Loan:
    All_Loans = {}
    id = {}
    dpd = {}
    product = {}
    client_id = {}

    def __init__ (self, id, product, dpd, client_id):
        self.id = id
        self.product = product
        self.dpd = dpd
        self.client_id = client_id

class Client:
    All_Clients = {}
    id = {}
    dpd = {}
    product = {}
    loan_id = {}

    def __init__ (self, id, product, dpd, loan_id):
        self.id = id
        self.product = product
        self.dpd = dpd
        self.loan_id = loan_id

loan_list_ds = []
loan_list_cl = []
loan_list_mini56 = []


tab_file=""
for i in range(1,len(sys.argv)):
    tab_file=tab_file+' '+sys.argv[i]
tab_file = tab_file.lstrip()
f = open(tab_file)


for row in f:
    row = row.replace('"','')
    l = row.split(',')
    if (l[3] == 'Test' or l[3] == 'Bankruptcy Initiated' or l[3] == 'Bankruptcy Confirmed' or l[0] == "LOAN ID"):
        #print (f'{l[0]} - {l[3]} : {l[15]}')
        continue
    else:
        r = Loan(l[0], l[4], l[15], l[13])
        Loan.All_Loans[l[0]] = r
        loan_list_ds.append(l[0])

f = open('/tmp/clients.csv')

for row in f:
    l = row.split(',')
    r = Client(l[0], l[11], l[6], l[2])
    if (l[11] == 'MINI 56' and int(l[6]) <= 10):
        loan_list_mini56.append(l[0])
    Client.All_Clients[l[2]] = r
    loan_list_cl.append(l[2])

print ('<table border=1><tr><td>DS</td><td>dpd</td><td>product</td><td>DMS</td><td></td></tr>')
for a in loan_list_ds:
    if a in loan_list_cl:
        continue
    else:
        loan = Loan.All_Loans.get(a)
        print (f'<tr><td><a href="https://ecocrm.ya.ecofin.io/ru/collection/show/{loan.id}/installment_pdl">{loan.id}</a></td>')
        print (f'<td>{loan.dpd}</td><td>{loan.product}</td><td><a href="https://dms.fin.dyninno.net/admin/client/{loan.client_id}">{loan.client_id}</a></td><td>ACTIVE</td></tr>')
print ('</table>')

print ('<table border=1><tr><td>DS</td><td>dpd</td><td>product</td><td>DMS</td><td></td>dpd</tr>')
for a in loan_list_ds:
    if a in loan_list_cl:
        loan = Loan.All_Loans.get(a)
        client = Client.All_Clients.get(a)
        if (loan.dpd == client.dpd):
            continue
        else:
            print (f'<tr><td><a href="https://ecocrm.ya.ecofin.io/ru/collection/show/{loan.id}/installment_pdl">{loan.id}</a></td>')
            print (f'<td>{loan.dpd}</td><td>{loan.product}</td><td><a href="https://dms.fin.dyninno.net/admin/client/{loan.client_id}">{loan.client_id}</a></td><td>{client.dpd}</td></tr>')
    else:
        continue
print ('</table>')

config_of_mess=load_workbook(filename='config.xlsx')
ws=config_of_mess['data']
for i in range(0, 1000):
    cell = 'I'+str(i+1)
    ws[cell] = None

a = 0
for i in range(len(loan_list_mini56)):
    a += 1
    cell = 'I'+str(a)
    ws[cell] = loan_list_mini56[i]

config_of_mess.save('config.xlsx')

file_mini56 = open('/tmp/mini56.html','w')
file_mini56.write('<table>')
for i in loan_list_mini56:
    file_mini56.write(f'<tr><td>{i}</td></tr>\n')
file_mini56.write('</table>')
file_mini56.close
