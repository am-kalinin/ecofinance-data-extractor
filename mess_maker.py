#!/usr/bin/python3

import csv, datetime, os.path, openpyxl

from openpyxl import load_workbook

f=str(datetime.datetime.now()).split()[0]+'.log'
used_loans=[]
if (os.path.exists(f)==True):
    f_log=open(f,'r')
    used_loans=f_log.read()
    used_loans=used_loans.split('\n')
    used_loans=list(filter(None,used_loans))
#    print (used_loans)
    f_log.close()
f_log=open(f,'a')
#print('used_loans len =', len(used_loans))
agentlist=[]
wb=load_workbook(filename='config.xlsx')
ws=wb['data']
def read_col(col):
    list1=[]
    for i in range(1,1000):
        ind=col+str(i)
        if (ws[ind].value is None): continue
        data=(str(ws[ind].value).rstrip())
        list1.append(data)
    return (list1)

agentlist=read_col('A')
statuses=read_col('B')
mini56list=read_col('I')

#print (statuses)
bucket1begin, bucket1end=ws['c2'].value.split()[0].split('-')
bucket2begin, bucket2end=ws['c3'].value.split()[0].split('-')
rcvryPDLZ1=int(ws['d2'].value.split('%')[0].split('.')[0])
rcvryPDLZ2=int(ws['d3'].value.split('%')[0].split('.')[0])
rcvryPDLR1=int(ws['e2'].value.split('%')[0].split('.')[0])
rcvryPDLR2=int(ws['e2'].value.split('%')[0].split('.')[0])
rcvryMini1=int(ws['f2'].value.split('%')[0].split('.')[0])
rcvryMini2=int(ws['f3'].value.split('%')[0].split('.')[0])
rcvryMini561=int(ws['g2'].value.split('%')[0].split('.')[0])
rcvryMini562=int(ws['g3'].value.split('%')[0].split('.')[0])
rcvry=[rcvryPDLZ1, rcvryPDLR1, rcvryPDLZ2, rcvryPDLR2, rcvryMini1, rcvryMini2, rcvryMini561, rcvryMini562]

action_loans=read_col('H')
def loan_check(string):
    if (len(used_loans)==0):
        return (True)
    else:
        for loan in used_loans:
#            print ("|"+loan.split(':')[0],"=", "["+string[0]+"]")
            if (string[0]==loan.split(':')[0]):
#                print ('False')
                return (False)
#                print (string[0],'=',loan.split(':')[0])
#    print (string[0],'!=',loan.split(':')[0])
#    row=string[0]+':'+reason+'\n'
#    used_loans.append(row)
    return (True)

today=datetime.datetime.now()
#print (today.strftime('%Y-%m-%d'))
ptpdate=[]
for d in range(-2,10):
    day=today+datetime.timedelta(days=d)
    ptpdate.append(day.strftime('%Y-%m-%d'))
print (ptpdate)
def ptpcheck(string):
    check=1
    for day in ptpdate:
        if (string==day):
#            print (f'!--({day})--!<br>')
            check=0
            continue
    return(check)

def statuscheck(row):
    check=1
    for status in statuses:
        if (status==row):
            check=0
            continue
    return(check)

def check_action(loan):
    if loan in action_loans:
        return ('<font color="red">Доступно списание процентов по акции</font>')
    else:
        return ('')

def htmlwriter(list):
#    list1=[]
    print ('<table border=1>')
    for row in list:
        print(f'<tr><td><a href="https://ecocrm.ya.ecofin.io/ru/collection/show/{row[0]}/installment_pdl">{row[0]}</a></td>')
        if (row[0] in mini56list):
            print ('<td>Mini56</td>')
        else:
            print (f'<td>{row[1]}</td>')
        print (f'<td>{row[2]}</td>'\
              f'<td>{row[3]}</td>'\
              f'<td>{row[4]}</td>'\
              f'<td>{row[5]}</td>'\
              f'<td>{check_action(row[0])}</td></tr>')
        string=row[0]+':'+row[5]+'\n'
        f_log.write(string)
#        list1.append(string)
    print ('</table>')
#    return (list1)

#print (int(rcvryPDLZ1))
#print (agentlist)

import sys

tab_file=""
for i in range(1,len(sys.argv)):
    tab_file=tab_file+' '+sys.argv[i]
tab_file = tab_file.lstrip()


PDL_Zero1=[]
PDL_Rep1=[]
Mini1=[]
Mini561=[]
Restr1=[]
PDL_Zero2=[]
PDL_Rep2=[]
Mini2=[]
Mini562=[]
Restr2=[]
new_table=[]

table=csv.reader(open(tab_file), delimiter=',')
a=b=c=d=e=0
for row in table:
        a=a+1
        if (row[10]=='---' or row[20]=='DELAYED AMOUNT'):
            b=b+1
            continue
        string = row[0], row[4], row[14], int(row[15]), int(row[20].split('.')[0])
        if (loan_check(string)==False):
            c=c+1
            continue
        if (statuscheck(row[12])==0):
            d=d+1
            string=row[0]+':'+row[12]+'\n'
            #used_loans.append(string)
            f_log.write(string)
            continue
        if (ptpcheck(row[9].split()[0])==0):
            e=e+1
            string=row[0]+':'+row[10].split()[0]+'\n'
            #used_loans.append(string)
            f_log.write(string)
            continue
        new_table.append(string)
print (f'{a} strings were operated, {b} have no personal task, {c} strings was already used loans, \
        {d} had unwanted statuses, {e} strings had actual ptp')
del table

import operator
new_table=sorted(new_table, key=operator.itemgetter(4), reverse=True)
bucket1, bucket2=[],[]

for row in new_table:
    if (row[3]>=int(bucket1begin) and row[3]<=int(bucket1end)):
        bucket1.append(row)
    elif (row[3]>=int(bucket2begin) and row[3]<=int(bucket2end)):
        bucket2.append(row)
del new_table

for row in bucket1:
    if (row[1]=='PDL Zero' or row[1]=='PDL New'):
            PDL_Zero1.append(row)
    elif (row[1]=='PDL Repeated'):
            PDL_Rep1.append(row)
    elif (row[1]=='Mini'):
        if (row[0] in mini56list):
            Mini561.append(row)
        else:
            Mini1.append(row)
    elif (row[1]=='Restructured 14'):
            Restr1.append(string)
del bucket1

for row in bucket2:
    if (row[1]=='PDL Zero' or row[1]=='PDL New'):
            PDL_Zero2.append(row)
    elif (row[1]=='PDL Repeated'):
            PDL_Rep2.append(row)
    elif (row[1]=='Mini'):
        if (row[0] in mini56list):
            Mini562.append(row)
        else:
            Mini2.append(row)
    elif (row[1]=='Restructured 14'):
            Restr2.append(string)
del bucket2

proportion = 20, 14, 12, 10, 8, 6, 4, 2

#print ('PDL Z', len(PDL_Zero2)//len(agentlist))
#print ('PDL R', len(PDL_Rep2)//len(agentlist))
#print ('Mini', len(Mini2)//len(agentlist))
#print ('Restr', len(Restr2)//len(agentlist))

#print (rcvry)
rcvry_sorted=sorted(rcvry)
#print (rcvry_sorted)

def countdefinder(recovery):
    n=0
    for i in rcvry_sorted:
        if (recovery==i):
#            print ('Recovery = ',recovery, 'Count of loans: ',proportion[n])
            return (proportion[n])
        n=n+1

final_table=[]


def fin_tab(recovery, product):
    o=c=0
    count=countdefinder(recovery)
    if (count>len(product)//len(agentlist)):
        count=len(product)//len(agentlist)
    while (c<count):
        i=0
        while (i<len(agentlist) and o<len(product)):
            string=product[o][0],product[o][1],product[o][2],product[o][3], product[o][4], agentlist[i]
            if (loan_check(string)==True):
                final_table.append(string)
                i=i+1
            else: 
                c=c-1
            o=o+1
        c=c+1
        
#print ('<br>PDL Zero 1 bucket: <br>')
fin_tab (rcvryPDLZ1,PDL_Zero1)
#print ('<br>PDL Zero 2 bucket: <br>')
fin_tab (rcvryPDLZ2,PDL_Zero2)
fin_tab (rcvryPDLR1,PDL_Rep1)
fin_tab (rcvryPDLR2,PDL_Rep2)
#print ('<br>PDL Mini 1 bucket: <br>')
fin_tab (rcvryMini1,Mini1)
#print ('<br>PDL Mini 2 bucket: <br>')
fin_tab (rcvryMini2,Mini2)
#print (len(final_table))
fin_tab(rcvryMini561,Mini561)
fin_tab(rcvryMini562,Mini562)

htmlwriter(final_table)

#f_log=open(f,'w')
#for i in final_table:
#    string=i[0]+':'+i[5]+'\n'
#    used_loans.append(string)
#for i in used_loans:
#    i=i+'\n'
#    f_log.write(i)
f_log.close()
#for i in used_loans:
#    agent_list.append([i[0],i[1]])
#config.save(filename=config)

