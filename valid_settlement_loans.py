#!/usr/bin/python3

import openpyxl

from openpyxl import load_workbook

wb=load_workbook(filename='config_settlement.xlsx')
ws=wb['Sheet1']
loans = []

def read_col(col):
    list1=[]
    for i in range(1,1000):
        ind=col+str(i)
        if (ws[ind].value is None): continue
        data=(str(ws[ind].value).rstrip())
        list1.append(data)
    return (list1)

term_days = int(ws['A2'].value)
valid_dpd = read_col('B')
configfile = 'config.xlsx'

filename='/tmp/settlement.csv'
f=open(filename)

for i in f:
    l = i.split(',')
    if (l[3] in valid_dpd and int(l[5]) >= term_days):
        loans.append(l[1])

print ('<table>') #<tr><td>PDLZ</td><td>PDLR</td><td>MINI</td><td>MINI56</td></tr>')
for a in range(len(loans)):
    print (f'<tr><td>{loans[a]}</td></td></tr>')
print ('</table>')

config_of_mess=load_workbook(filename=configfile)
ws=config_of_mess['data']

for i in range(0, 1000):
    cell = 'H'+str(i+1)
    ws[cell] = None

a = 0
for i in range(len(loans)):
    a += 1
    cell = 'H'+str(a)
    ws[cell] = loans[i]

config_of_mess.save(configfile)
